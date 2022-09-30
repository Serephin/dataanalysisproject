#!/usr/bin/env python
# coding: utf-8

# Code bellow reads Statistics on Death appeared due to cancer. This code is able to read excel file, get specific cell value and diagram about death cases in the period given by the user
# 

# In[21]:


import pandas as pd


# In[22]:


from openpyxl.workbook import Workbook


# In[23]:


df_excel=pd.read_excel('CancerStati.xlsx')


# In[24]:


df_excel.columns = df_excel.columns.str.strip('_ †')
print(df_excel)


# Bar chart:)

# In[25]:


print("Choose type of cancer that caused death cases: ")
for i in df_excel.columns:
    print(i)
choice=input()
print("Now write the starting and ending years of the period you want to see")
start=input()
end=input()
import matplotlib.pyplot as plt

i=0
n=0
list=[]
list2=[]

while i<83  :

    
    list2.append(str(df_excel.iloc[i,0]))
    n=n+1
    i=i+1
list=df_excel[choice]  
dict={}
i=0
k=i
while i<len(list) :
       
            dict[list2[k]]=list[i]
            
            
            i=i+1
            k=k+1
print(dict)
st=list2.index(start)
ed=list2.index(end)
fin=ed+1
print(fin)
revlist=list[st:fin]
revlist2=list2[st:fin]

plt.figure(figsize=(20,10))

plt.bar(revlist2,revlist)
plt.show()


# In[10]:


print(df_excel)


# In[26]:


df_excel["Total"]=df_excel['Stomach']+df_excel['Colon and Rectum']+df_excel['Pancreas']+df_excel['Lung and Bronchus']+df_excel['Breast']+df_excel['Uterus']+df_excel['Liver']


# In[27]:


df_excel.columns = df_excel.columns.str.strip('_ ')


# In[28]:


print(df_excel)


# Code below opens, reads data on the Credit Limit (Limit) based on Income, Age, Gender, Ethnicity and other factors. It fits a multiple linear model to predict the Credit Limit (Limit) based on multiple criterias of a person. It creates a regression scutterplot.  

# In[ ]:


from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import matplotlib
import numpy as np
from sklearn import linear_model


# In[3]:


import pandas as pd
#wb1=load_workbook(r'D:\\SarvinaWeek14Stat.xlsx')
wb1 = pd.read_excel (r'D:\\SarvinaWeek14Stat.xlsx',sheet_name=["credit"]) 
df=pd.concat(wb1[frame] for frame in wb1.keys())
df.columns = df.columns.str.strip('_  ')


# In[31]:


df


# In[4]:


df
median=df.Income.median()
print(median)
#newsheet=wb1.create_sheet("Sam")


# In[5]:


df.Income=df.Income.fillna(median)


# In[6]:


reg=linear_model.LinearRegression()
reg.fit(df[['Income', 'Age']], df.Limit)


# In[7]:


reg.coef_


# In[38]:


reg.intercept_


# In[39]:


reg.predict([[15000 , 35]])


# In[ ]:


print(df['Income'])
print("Choose numbers of columns with factors to make a regression and predictions, Limit is the data you are making prediction on so ignore it")
ndf=df['Limit']
for i in df.columns:
    print(i)
n=int(input())
print("List that  factors")
i=0
while i<n:
    inp=input()
    if inp=="Gender" or inp=="Ethnicity" or inp=="Student" or inp=="Married":
        dummies=pd.get_dummies(df[inp])
        ndf=pd.concat([ndf,dummies], axis='columns')
    else:
        ndf=pd.concat([ndf,df[inp]], axis='columns')
    i+=1
ndf


# In[ ]:


dummies=pd.get_dummies(df[inp])


# In[152]:


unit=pd.concat([df,dummies], axis='columns')
unit


# In[95]:





# In[41]:


X=ndf.drop(['Limit'], axis='columns')


# In[42]:


Y=ndf['Limit']


# In[43]:


import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split 
X_test, X_train, Y_test, Y_train=train_test_split(X,Y, test_size=0.3, random_state=0)


# In[44]:


final=linear_model.LinearRegression()

final.fit(X_train,Y_train)
final


# In[ ]:





# In[45]:


pred=final.predict(X_test)


# In[46]:


import matplotlib.pyplot as plt
plt.scatter(Y_test,pred)


# In[15]:


import sqlite3
db=sqlite3.connect('new.db')
cur=db.cursor()
 
for row in cur.execute(""" SElECT * from testik """):
     print (row)


# In[1]:


print("Hii")


# Winter is Coming. Let's load the dataset ASAP!
# If you haven't heard of Game of Thrones, then you must be really good at hiding. Game of Thrones is the hugely popular television series by HBO based on the (also) hugely popular book series A Song of Ice and Fire by George R.R. Martin. In this notebook, we will analyze the co-occurrence network of the characters in the Game of Thrones books. Here, two characters are considered to co-occur if their names appear in the vicinity of 15 words from one another in the books.
# 
# 
# 
# This dataset constitutes a network and is given as a text file describing the edges between characters, with some attributes attached to each edge. Let's start by loading in the data for the first book A Game of Thrones and inspect it.
# Зима приближается. Давайте загрузим набор данных как можно скорее!
# Если вы не слышали об «Игре престолов», то вы, должно быть, очень хорошо умеете прятаться. «Игра престолов» — чрезвычайно популярный телесериал HBO, основанный на (также) чрезвычайно популярной серии книг «Песнь льда и пламени» Джорджа Р. Р. Мартина. В этой записной книжке мы проанализируем сеть совпадений персонажей в книгах «Игра престолов». Здесь считается, что два персонажа встречаются одновременно, если их имена появляются в книгах на расстоянии около 15 слов друг от друга.
# 
# 
# 
# Этот набор данных представляет собой сеть и предоставляется в виде текстового файла, описывающего границы между символами, с некоторыми атрибутами, прикрепленными к каждому краю. Давайте начнем с загрузки данных для первой книги «Игра престолов» и проверим их.
# 

# In[8]:


import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import matplotlib
import numpy as np
got = pd.read_excel (r'GameofThrones.xlsx') 
got
target=got.loc[got['weight']<=15]
target=target.head(15)
targ=got['Target'].head(15)

import matplotlib.pyplot as plt
plt.bar(targ,target['weight'])
plt.xlabel("Pair Of Names")
plt.ylabel("Distance in words")
plt.xticks(rotation=90)

plt.show






# In[ ]:





# In[2]:





# In[ ]:





# In[ ]:





# In[ ]:




